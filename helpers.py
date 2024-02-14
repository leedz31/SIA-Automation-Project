import os
import pandas as pd
from openpyxl import Workbook, load_workbook


def convert_to_xlsx(src_files):
    xlsx_files = []
    for source_file in src_files:
        read_file = pd.read_csv(source_file)
        converted_file_path = source_file[:-4] + '.xlsx'
        read_file.to_excel(converted_file_path, index = None, header = True)
        xlsx_files.append(converted_file_path)
    return xlsx_files;

def combine_workbooks(xlsx_files, dest_file):
    # this function takes in an array of xlsx files and a destination file path
    combined_wb = pd.DataFrame()

    # concat all the worksheets into the combined_wb dataframe
    for source_file in xlsx_files:
        df = pd.read_excel(source_file, engine = 'openpyxl')
        combined_wb = pd.concat([combined_wb,df], ignore_index = True)

    # if the destination file does not exist, create one
    if not os.path.exists(dest_file):
        combined_wb.to_excel(dest_file, sheet_name='Combined Data', index=False, engine = 'openpyxl')
    else:
        with pd.ExcelWriter(dest_file, engine='openpyxl', mode='a' if os.path.exists(dest_file) else 'w') as writer:
            combined_wb.to_excel(writer, sheet_name='Combined Data', index=False)
    
    return 

def copy_data(wb, new_sheet_name):
    # the dest file path will be added as the wb parameter
    workbook = load_workbook(wb)
    src_sheet = workbook.active
    # create a new worksheet and name it duplicates or phished
    new_sheet = workbook.create_sheet(new_sheet_name)
    # paste the copied data onto the new worksheet
    for row in src_sheet.iter_rows(values_only = True):
        new_sheet.append(row)
    workbook.save(wb)
    return

def filter_data_to_new_ws(wb, ws, new_sheet_name, filter_conditions):
    df = pd.read_excel(wb, sheet_name = ws)

    # df now contains the filtered data
    for col, cond in filter_conditions:
        df = df[df[col].apply(cond)]

    workbook = load_workbook(wb)

    if new_sheet_name in workbook.sheetnames:
        std = workbook[new_sheet_name]
        workbook.remove(std)
    workbook.save(wb)

    with pd.ExcelWriter(wb, engine='openpyxl', mode='a' if os.path.exists(wb) else 'w') as writer:
        df.to_excel(writer, sheet_name= new_sheet_name, index=False)
    
    return 

    # def cond1(x):
    #     return x>50
    filter_conditions = [
        ('Column 1', cond1),
        ('Column 2', cond2)
    ]
    filtered_data = filter_data

def add_column(wb, ws, new_col_name):
    # read the ws into a dataframe df
    df = pd.read_excel(wb, sheet_name = ws)
    df[new_col_name] = df.groupby('Email')['Email'].transform('count')

    with pd.ExcelWriter(wb, engine = 'openpyxl', mode = 'a') as writer:
        # get the index of the worksheet we want to overwrite
        if ws in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(ws)
            writer.book.remove(writer.book.worksheets[idx])

        df.to_excel(writer, sheet_name = ws, index = False)




    
